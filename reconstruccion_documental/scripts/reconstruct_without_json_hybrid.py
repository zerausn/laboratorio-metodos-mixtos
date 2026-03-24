from __future__ import annotations

import argparse
import json
import re
import sys
import textwrap
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

import cv2
import fitz
import numpy as np
import pandas as pd
import pytesseract
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.pdfgen import canvas

SCRIPTS_DIR = Path(__file__).resolve().parent
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from docai_pdf_pipeline import (
    configure_tesseract,
    preprocess_for_ocr,
    render_page,
    resolve_tesseract,
    rotate_image,
    score_ocr,
)
from project_paths import OUTPUTS_DIR, PDF_PATH, TESSDATA_DIR
from reconstruct_clean_document import normalize_spaces, ocr_clean_page

DEFAULT_PAGE3_HEADER = [
    "SANTIAGO DE CALI",
    "DEPARTAMENTO ADMINISTRATIVO DE HACIENDA",
    "SUBDIRECCION DE FINANZAS PUBLICAS",
    "SUBPROCESO DE SEGUIMIENTO, ANALISIS Y CONTROL PRESUPUESTAL",
    "PRESUPUESTO Y EJECUCION DE INGRESOS SEPTIEMBRE 2025",
    "ADMINISTRACION CENTRAL",
]


@dataclass
class HybridCell:
    row_index: int
    col_index: int
    pdf_text: str
    ocr_text: str
    final_text: str
    selected_source: str


@dataclass
class HybridPage:
    absolute_page: int
    rotation: int
    mode: str
    header_lines: list[str]
    row_boundaries: list[int]
    col_boundaries: list[int]
    cells: list[HybridCell]
    preview: str
    line_fallback: list[dict[str, Any]]


def clean_text(text: str) -> str:
    for old, new in {"“": '"', "”": '"', "’": "'", "‘": "'", "—": "-", "–": "-", "•": "-", "Â": ""}.items():
        text = (text or "").replace(old, new)
    return normalize_spaces(text)


def cluster(values: list[int], gap: int) -> list[int]:
    if not values:
        return []
    values = sorted(values)
    groups = [[values[0]]]
    for value in values[1:]:
        if value - groups[-1][-1] <= gap:
            groups[-1].append(value)
        else:
            groups.append([value])
    return [int(round(sum(group) / len(group))) for group in groups]


def transform_bbox(page: fitz.Page, zoom: float, rotation: int, bbox: tuple[float, float, float, float]) -> tuple[float, float, float, float]:
    x0, y0, x1, y1 = bbox
    width = page.rect.width * zoom
    height = page.rect.height * zoom
    if rotation == 90:
        return height - y1 * zoom, x0 * zoom, height - y0 * zoom, x1 * zoom
    if rotation == 270:
        return y0 * zoom, width - x1 * zoom, y1 * zoom, width - x0 * zoom
    return x0 * zoom, y0 * zoom, x1 * zoom, y1 * zoom


def choose_rotation(page: fitz.Page) -> tuple[int, np.ndarray]:
    base = render_page(page, zoom=4.0)
    candidates: list[tuple[float, int, np.ndarray]] = []
    for rotation in (0, 90, 270):
        rotated = rotate_image(base, rotation)
        processed = preprocess_for_ocr(rotated)
        data = pytesseract.image_to_data(processed, lang="spa+eng", config="--psm 6", output_type=pytesseract.Output.DICT)
        avg_conf, weird_ratio, word_count, score = score_ocr(data)
        bbox = detect_table_bbox(rotated)
        bonus = 0
        if bbox is not None:
            bonus = 20
            if bbox[2] > bbox[3]:
                bonus += 25
        bias = 5 if rotation == 90 else 0
        final = score + bonus + bias - (weird_ratio * 10 if avg_conf < 20 and word_count < 20 else 0)
        candidates.append((final, rotation, rotated))
    _, rotation, image = max(candidates, key=lambda item: item[0])
    return rotation, image


def detect_table_bbox(image: np.ndarray) -> tuple[int, int, int, int] | None:
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    _, inv = cv2.threshold(gray, 200, 255, cv2.THRESH_BINARY_INV)
    contours, _ = cv2.findContours(inv, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    image_area = image.shape[0] * image.shape[1]
    candidates = []
    for contour in contours:
        x, y, w, h = cv2.boundingRect(contour)
        area = w * h
        if area < image_area * 0.08 or area > image_area * 0.97:
            continue
        if w < image.shape[1] * 0.35 or h < image.shape[0] * 0.25:
            continue
        candidates.append((x, y, w, h))
    return max(candidates, key=lambda item: item[2] * item[3]) if candidates else None


def detect_grid(crop: np.ndarray) -> tuple[list[int], list[int], np.ndarray]:
    gray = cv2.cvtColor(crop, cv2.COLOR_BGR2GRAY)
    _, binary = cv2.threshold(255 - gray, 120, 255, cv2.THRESH_BINARY)
    horizontal = cv2.morphologyEx(binary, cv2.MORPH_OPEN, cv2.getStructuringElement(cv2.MORPH_RECT, (max(40, crop.shape[1] // 35), 1)))
    vertical = cv2.morphologyEx(binary, cv2.MORPH_OPEN, cv2.getStructuringElement(cv2.MORPH_RECT, (1, max(40, crop.shape[0] // 35))))
    row_positions = cluster([i for i, value in enumerate(horizontal.sum(axis=1)) if value > 255 * crop.shape[1] * 0.23], max(3, crop.shape[0] // 900))
    col_positions = cluster([i for i, value in enumerate(vertical.sum(axis=0)) if value > 255 * crop.shape[0] * 0.23], max(3, crop.shape[1] // 900))
    return row_positions, col_positions, cv2.bitwise_or(horizontal, vertical)


def extract_header_lines(page_number: int, image: np.ndarray, bbox: tuple[int, int, int, int]) -> list[str]:
    if page_number == 3:
        return DEFAULT_PAGE3_HEADER
    _, y, xw, _ = bbox
    crop = image[max(0, y - 320) : max(1, y - 10), : min(image.shape[1], max(500, xw + 150))]
    if crop.size == 0:
        return [f"Reconstruccion sin JSON - pagina {page_number}"]
    text = pytesseract.image_to_string(preprocess_for_ocr(crop), lang="spa+eng", config="--psm 11")
    lines = []
    for raw_line in text.splitlines():
        line = clean_text(raw_line)
        if len(line) < 6 or sum(ch.isalpha() for ch in line) < 4:
            continue
        lines.append(line)
    return lines[:6] or [f"Reconstruccion sin JSON - pagina {page_number}"]


def words_to_cell_text(words: list[tuple[float, float, str]], x0: int, y0: int, x1: int, y1: int) -> str:
    items = sorted([(cy, cx, text) for cx, cy, text in words if x0 <= cx < x1 and y0 <= cy < y1], key=lambda item: (item[0], item[1]))
    lines: list[list[tuple[float, float, str]]] = []
    for item in items:
        if not lines or abs(item[0] - lines[-1][-1][0]) > 12:
            lines.append([item])
        else:
            lines[-1].append(item)
    return clean_text(" ".join(" ".join(part[2] for part in line) for line in lines))


def column_mode(col_index: int, total_cols: int) -> str:
    if col_index == 0:
        return "id"
    if col_index in {1, 3}:
        return "code"
    if col_index >= max(0, total_cols - 4):
        return "numeric"
    return "text"


def normalize_candidate(text: str, mode: str) -> str:
    text = clean_text(text)
    if not text:
        return ""
    if mode == "id":
        return re.sub(r"[^0-9]", "", text)
    if mode == "code":
        return clean_text(re.sub(r"[^0-9A-Za-z.\- ]", "", text.replace(",", ".")))
    if mode == "numeric":
        return clean_text(re.sub(r"[^0-9.,%\-() ]", "", text))
    return clean_text(re.sub(r"\s+", " ", text.replace("|", " "))).strip(" -")


def score_candidate(text: str, mode: str) -> float:
    if not text:
        return -100.0
    digits = sum(ch.isdigit() for ch in text)
    letters = sum(ch.isalpha() for ch in text)
    spaces = text.count(" ")
    dots = text.count(".")
    commas = text.count(",")
    weird = sum(not (ch.isalnum() or ch in " .,%-()/") for ch in text)
    if mode == "id":
        return digits * 2.5 - letters * 1.5 - weird * 4 - spaces
    if mode == "code":
        return digits * 1.6 + dots * 1.5 - letters * 0.4 - weird * 3
    if mode == "numeric":
        return digits * 1.8 + dots + commas - letters * 2.2 - weird * 4
    return letters * 1.2 + digits * 0.2 + spaces * 0.3 - weird * 2


def ocr_cell(cell: np.ndarray, mode: str) -> str:
    if cell.size == 0:
        return ""
    resized = cv2.resize(cell, None, fx=6 if cell.shape[1] < 80 else 4, fy=6 if cell.shape[0] < 24 else 4, interpolation=cv2.INTER_CUBIC)
    thresholded = cv2.threshold(resized, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)[1]
    if mode == "numeric":
        config = "--psm 7 -c tessedit_char_whitelist=0123456789.,%-()"
    elif mode in {"id", "code"}:
        config = "--psm 7 -c tessedit_char_whitelist=0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz.-"
    else:
        config = "--psm 6"
    return clean_text(pytesseract.image_to_string(thresholded, lang="spa+eng", config=config))


def choose_text(row_index: int, col_index: int, total_cols: int, pdf_text: str, ocr_text: str) -> tuple[str, str]:
    mode = column_mode(col_index, total_cols)
    pdf_norm = normalize_candidate(pdf_text, mode)
    ocr_norm = normalize_candidate(ocr_text, mode)
    pdf_score = score_candidate(pdf_norm, mode) + (1.2 if row_index >= 6 and mode == "numeric" else 0)
    ocr_score = score_candidate(ocr_norm, mode) + (1.5 if row_index < 6 and mode == "text" else 0)
    if ocr_score > pdf_score + 1:
        return ocr_norm, "ocr"
    if pdf_norm:
        return pdf_norm, "pdf_text"
    return ocr_norm, "ocr"


def matrix_from_cells(cells: list[HybridCell], total_rows: int, total_cols: int) -> list[list[str]]:
    rows = [["" for _ in range(total_cols)] for _ in range(total_rows)]
    for cell in cells:
        rows[cell.row_index][cell.col_index] = cell.final_text
    for row in rows:
        detail_hint = any(keyword in clean_text(" ".join(row[2:5])).lower() for keyword in ("sobretasa", "impuesto", "vigencia", "predial", "gasolina", "retelca"))
        if detail_hint and not row[0] and row[3]:
            row[0] = "4131"
    return rows


def preview_from_rows(rows: list[list[str]]) -> str:
    chunks = []
    for row in rows[:6]:
        text = " | ".join(cell for cell in row if cell)
        if text:
            chunks.append(text)
    return clean_text(" ".join(chunks))[:180]


def reconstruct_table_page(page: fitz.Page, page_number: int) -> HybridPage | None:
    rotation, image = choose_rotation(page)
    bbox = detect_table_bbox(image)
    if bbox is None:
        return None
    x, y, w, h = bbox
    crop = image[y : y + h, x : x + w]
    row_boundaries, col_boundaries, grid_lines = detect_grid(crop)
    if len(row_boundaries) < 4 or len(col_boundaries) < 4:
        return None

    gray_crop = cv2.cvtColor(crop, cv2.COLOR_BGR2GRAY)
    line_removed = cv2.bitwise_or(gray_crop, grid_lines)
    transformed_words: list[tuple[float, float, str]] = []
    for word in page.get_text("words"):
        wx0, wy0, wx1, wy1 = transform_bbox(page, zoom=4.0, rotation=rotation, bbox=word[:4])
        cx = (wx0 + wx1) / 2
        cy = (wy0 + wy1) / 2
        if x <= cx <= x + w and y <= cy <= y + h:
            transformed_words.append((cx - x, cy - y, str(word[4])))

    total_rows = len(row_boundaries) - 1
    total_cols = len(col_boundaries) - 1
    cells: list[HybridCell] = []
    for row_index in range(total_rows):
        for col_index in range(total_cols):
            x0, x1 = col_boundaries[col_index] + 2, col_boundaries[col_index + 1] - 2
            y0, y1 = row_boundaries[row_index] + 2, row_boundaries[row_index + 1] - 2
            if x1 <= x0 or y1 <= y0:
                continue
            pdf_text = words_to_cell_text(transformed_words, x0, y0, x1, y1)
            ocr_text = ocr_cell(line_removed[y0:y1, x0:x1], column_mode(col_index, total_cols))
            final_text, source = choose_text(row_index, col_index, total_cols, pdf_text, ocr_text)
            cells.append(HybridCell(row_index, col_index, pdf_text, ocr_text, final_text, source))

    rows = matrix_from_cells(cells, total_rows, total_cols)
    return HybridPage(
        absolute_page=page_number,
        rotation=rotation,
        mode="table_grid",
        header_lines=extract_header_lines(page_number, image, bbox),
        row_boundaries=row_boundaries,
        col_boundaries=col_boundaries,
        cells=cells,
        preview=preview_from_rows(rows),
        line_fallback=[],
    )


def reconstruct_page(page: fitz.Page, page_number: int) -> HybridPage:
    table_page = reconstruct_table_page(page, page_number)
    if table_page is not None:
        return table_page
    fallback = ocr_clean_page(page, page_number)
    return HybridPage(
        absolute_page=page_number,
        rotation=0,
        mode="ocr_lines",
        header_lines=[f"Reconstruccion OCR sin JSON - pagina {page_number}"],
        row_boundaries=[],
        col_boundaries=[],
        cells=[],
        preview=fallback.preview,
        line_fallback=[{"text": line.text, "x0": line.x0, "y0": line.y0, "x1": line.x1, "y1": line.y1} for line in fallback.lines],
    )


def write_excel(output_xlsx: Path, pages: list[HybridPage]) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    summary = workbook.create_sheet("summary")
    summary.append(["absolute_page", "mode", "rotation", "rows", "cols", "preview"])
    for page in pages:
        summary.append([page.absolute_page, page.mode, page.rotation, max(0, len(page.row_boundaries) - 1), max(0, len(page.col_boundaries) - 1), page.preview])
        sheet = workbook.create_sheet(f"p{page.absolute_page:03d}")
        cursor = 1
        for line in page.header_lines:
            sheet.cell(row=cursor, column=1, value=line)
            cursor += 1
        cursor += 1
        if page.mode == "table_grid":
            rows = matrix_from_cells(page.cells, len(page.row_boundaries) - 1, len(page.col_boundaries) - 1)
            widths = [page.col_boundaries[index + 1] - page.col_boundaries[index] for index in range(len(page.col_boundaries) - 1)]
            for col_index, width in enumerate(widths, start=1):
                sheet.column_dimensions[get_column_letter(col_index)].width = max(8, min(55, width / 10))
            for row in rows:
                for col_index, value in enumerate(row, start=1):
                    sheet.cell(row=cursor, column=col_index, value=value)
                cursor += 1
        else:
            sheet.append(["line_index", "text", "x0", "y0", "x1", "y1"])
            cursor += 1
            for index, item in enumerate(page.line_fallback, start=1):
                sheet.append([index, item["text"], item["x0"], item["y0"], item["x1"], item["y1"]])
        sheet.freeze_panes = "A2"
    workbook.save(output_xlsx)


def split_cell_text(text: str, width: float, font_size: float) -> list[str]:
    text = clean_text(text)
    if not text:
        return []
    wrap = max(4, int(width / max(font_size * 0.58, 1)))
    return textwrap.wrap(text, width=wrap, break_long_words=False, break_on_hyphens=False)[:4] or [text]


def draw_page(pdf: canvas.Canvas, page: HybridPage) -> None:
    page_width, page_height = landscape(letter)
    margin, footer = 20, 14
    pdf.setPageSize((page_width, page_height))
    pdf.setFillColor(colors.white)
    pdf.rect(0, 0, page_width, page_height, fill=1, stroke=0)
    pdf.setFillColor(colors.black)
    y_cursor = page_height - margin
    pdf.setFont("Helvetica", 8)
    for line in page.header_lines[:6]:
        pdf.drawString(margin, y_cursor, line)
        y_cursor -= 10

    if page.mode == "table_grid":
        rows = matrix_from_cells(page.cells, len(page.row_boundaries) - 1, len(page.col_boundaries) - 1)
        row_heights = [max(8, page.row_boundaries[index + 1] - page.row_boundaries[index]) for index in range(len(page.row_boundaries) - 1)]
        col_widths = [max(20, page.col_boundaries[index + 1] - page.col_boundaries[index]) for index in range(len(page.col_boundaries) - 1)]
        available_height = max(80, y_cursor - margin - footer - 8)
        available_width = page_width - (margin * 2)
        row_scale = available_height / max(sum(row_heights), 1)
        col_scale = available_width / max(sum(col_widths), 1)
        scaled_rows = [value * row_scale for value in row_heights]
        scaled_cols = [value * col_scale for value in col_widths]
        current_y = y_cursor - 8
        for row_index, row in enumerate(rows):
            current_x = margin
            row_height = scaled_rows[row_index]
            for col_index, value in enumerate(row):
                col_width = scaled_cols[col_index]
                pdf.rect(current_x, current_y - row_height, col_width, row_height, fill=0, stroke=1)
                for offset, line in enumerate(split_cell_text(value, col_width - 4, 5)):
                    text_y = current_y - 6 - (offset * 5.5)
                    if text_y <= current_y - row_height + 2:
                        break
                    pdf.setFont("Helvetica", 5)
                    pdf.drawString(current_x + 2, text_y, line)
                current_x += col_width
            current_y -= row_height
            if current_y <= margin + footer:
                break
    else:
        pdf.setFont("Helvetica", 5)
        current_y = y_cursor
        for item in page.line_fallback:
            current_y -= 7
            if current_y <= margin + footer:
                break
            pdf.drawString(margin, current_y, clean_text(item["text"])[:180])

    pdf.setFont("Helvetica", 8)
    pdf.drawRightString(page_width - margin, footer / 2, f"PDF page {page.absolute_page}")
    pdf.showPage()


def write_pdf(output_pdf: Path, pages: list[HybridPage]) -> None:
    pdf = canvas.Canvas(str(output_pdf))
    for page in pages:
        draw_page(pdf, page)
    pdf.save()


def write_manifest(output_json: Path, pages: list[HybridPage]) -> None:
    output_json.write_text(json.dumps({"pipeline": "reconstruct_without_json_hybrid", "pages": [asdict(page) for page in pages]}, indent=2, ensure_ascii=False), encoding="utf-8")


def parse_pages(raw: str | None, page_count: int) -> list[int]:
    if not raw:
        return list(range(1, page_count + 1))
    selected: set[int] = set()
    for chunk in raw.split(","):
        chunk = chunk.strip()
        if not chunk:
            continue
        if "-" in chunk:
            start, end = chunk.split("-", 1)
            selected.update(range(int(start), int(end) + 1))
        else:
            selected.add(int(chunk))
    return sorted(page for page in selected if 1 <= page <= page_count)


def run_pipeline(pdf_path: Path, output_dir: Path, pages_arg: str | None, tesseract_cmd: str | None, tessdata_dir: Path | None, artifact_prefix: str) -> list[HybridPage]:
    output_dir.mkdir(parents=True, exist_ok=True)
    configure_tesseract(resolve_tesseract(tesseract_cmd), tessdata_dir)
    pdf_doc = fitz.open(str(pdf_path))
    pages = [reconstruct_page(pdf_doc.load_page(page_number - 1), page_number) for page_number in parse_pages(pages_arg, pdf_doc.page_count)]
    write_pdf(output_dir / f"{artifact_prefix}.pdf", pages)
    write_excel(output_dir / f"{artifact_prefix}.xlsx", pages)
    write_manifest(output_dir / f"{artifact_prefix}.json", pages)
    return pages


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Reconstruct pages without JSON using PDF text + OCR + grid detection.")
    parser.add_argument("--pdf", default=str(PDF_PATH), help="Absolute path to the original PDF.")
    parser.add_argument("--output-dir", default=str(OUTPUTS_DIR / "no_json_hybrid_full"), help="Directory for generated outputs.")
    parser.add_argument("--pages", default=None, help="Pages to process, e.g. 3 or 1-5,9.")
    parser.add_argument("--tesseract-cmd", default=None, help="Optional explicit path to tesseract.exe.")
    parser.add_argument("--tessdata-dir", default=str(TESSDATA_DIR), help="Optional tessdata directory.")
    parser.add_argument("--artifact-prefix", default="document_reconstructed_no_json_hybrid", help="Base name for the generated outputs.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    pages = run_pipeline(Path(args.pdf), Path(args.output_dir), args.pages, args.tesseract_cmd, Path(args.tessdata_dir) if args.tessdata_dir else None, args.artifact_prefix)
    print(f"Done. Reconstructed {len(pages)} page(s) with the no-JSON hybrid pipeline.")


if __name__ == "__main__":
    main()
